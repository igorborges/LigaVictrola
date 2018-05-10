import json, requests

import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
# -*- coding: utf-8 -*-
import sys
import time

reload(sys)
sys.setdefaultencoding('utf8')
import getpass
import urllib3

urllib3.disable_warnings()


# ----------------------------------------------------------------------------------- LOGIN
def login(user, password):
    body = {
        "payload":
            {
                "email": user,
                "password": password,
                "serviceId": 438

            },
        "captcha": ""
    }
    header = {'Content-Type': 'application/json; charset=UTF-8'}
    response = requests.post("https://login.globo.com/api/authentication", data=json.dumps(body), headers=header,
                             verify=False)
    print response.json()['userMessage']
    return response.json()['glbId']


# ----------------------------------------------------------------------------------- CRIA LIGA
def criaLiga(token, nomeDaLiga, quantidadeDeTimes):
    body = {
        "tipo": "F",
        "mata_mata": True,
        "quantidade_times": int(quantidadeDeTimes),
        "dataInicioRodada": "",
        "dataFimRodada": "",
        "descricao": "Cadastro em para futuros convites: https://goo.gl/forms/i2rXZ49urLSAP5dQ2 Ou se garanta fazendo um investimento de R$1,00 por liga. Zap (64)99979-8012",
        "nome": nomeDaLiga,
        "fim_rodada": None,
        "tipo_trofeu": 1,
        "cor_trofeu": 1
    }
    header = {'x-glb-token': token}
    response = requests.post("https://api.cartolafc.globo.com/auth/liga", data=json.dumps(body), headers=header,
                             verify=False)
    print response.json()['mensagem']
    if response.status_code == 201:
        return [201, response.json()['slug']]


# ----------------------------------------------------------------------------------- ENVIA CONVITES
def enviaConvites(token, liga, times_para_convidar):
    url = "https://api.cartolafc.globo.com/auth/liga/" + liga + "/convidar"
    headers = {'x-glb-token': token}
    response = requests.request("POST", url, data=json.dumps(times_para_convidar), headers=headers, verify=False)
    print(response.json()['mensagem'])


# ----------------------------------------------------------------------------------- CHECA TIMES CONVIDADOS
def checaTimesConvidados(token, liga, timesQueJaAceitaram, timesJaConvidados):
    url = "https://api.cartolafc.globo.com/auth/liga/" + liga
    headers = {
        'x-glb-token': token,
        'Cache-Control': "no-cache",
        'Postman-Token': "64148415-e0f3-472a-ba62-2987db10f1cf"
    }
    response = requests.request("GET", url, headers=headers, verify=False)

    if response.status_code == 200:
        try:
            times_convidados = response.json()['convites_enviados']
            # print "\nTimes que ainda nao aceitaram o convite:"
            for time in times_convidados:
                # print " - " + time['time']['nome']
                timesJaConvidados.append(time['time']['nome'])
        except:
            pass
            # print "Todos os times ja aceitaram o convite\n"

        times_garantidos = response.json()['times']
        # print "Times que ja aceitaram o convite:"
        for time in times_garantidos:
            # print " - " + time['nome']
            timesQueJaAceitaram.append(time['nome'])
    else:
        print response


# ----------------------------------------------------------------------------------- APAGA A LIGA (so usar pra teste)
def apagaLiga(token, liga):
    header = {'x-glb-token': token}
    response = requests.delete("https://api.cartolafc.globo.com/auth/liga/" + liga, headers=header, verify=False)
    print ""
    print response.json()['mensagem']


# ----------------------------------------------------------------------------------- BUSCA TIMES PARA CONVITE
def buscaTimes(deveSerConvidado, workbookPath, numeroDaLiga):
    wb = openpyxl.load_workbook(filename=workbookPath)  # abre o arquivo
    ranking = wb['Para convites']

    col = -1
    for i in xrange(5, 100):
        try:
            if ranking.cell(2, i).value.lower() == numeroDaLiga.lower():
                col = i
                break
        except:
            pass

    if col == -1:
        print "Liga nao encontrada na planilha, confira o numero e tente novamente."
        return []

    times = []  # cria um array vazio
    lin = 3
    time = ranking.cell(lin, 2).value  # pega o primeiro time do ranking
    while (time != None):  # enquanto o campo nao for vazio
        time = time.lower().replace(" ", "-").replace(".",
                                                      "-")  # pega o time, passa pra minusculo e tira os caracteres especiais e espacos
        # print time
        if time[time.__len__() - 1] == '-':  # se o time terminar com "-", remove o "-" da ultima posicao
            time = time[:-1]
        # response = requests.get('https://api.cartolafc.globo.com/time/slug/' + time, verify=False) # faz a requisicao pra saber se o time existe
        # print time + " -> " + str(response.json()['time']['nome_cartola']) # imprime o time e o nome do cartoleiro
        try:
            if ranking.cell(lin, col).value.lower() == deveSerConvidado or deveSerConvidado == 'a':
                times.append(time)  # adiciona o time na lista
        except:
            pass
        # ranking.cell(lin, 5).value = ranking.cell(lin, 5).value + '*'
        lin += 1
        time = ranking.cell(lin, 2).value  # pega proximo time da coluna

    # wb.save("teste1.xlsx")
    return times  # imprime a lista
    # ws = wb.worksheets[0]
    # c = ["Third", 40, 20, 35, 25, 20, 35]

    # for i in range(len(c)):
    #     ws.cell(row=i + 1, column=3).value = c[i]

    # lc = openpyxl.chart.LineChart()
    # lc.title = "Three Lines Chart"
    # data = openpyxl.chart.Reference(ws, min_col=1, min_row=1, max_col=3, max_row=len(c))
    # lc.add_data(data, titles_from_data=True)

    # ws.add_chart(lc, "D1")


# ----------------------------------------------------------------------------------- BUSCA TIMES POR RANKING
def buscaTimesPorRanking(numeroDeTimesConvidados, workbookPath):
    wb = openpyxl.load_workbook(filename=workbookPath)  # abre o arquivo
    ranking = wb.worksheets[0]  # pega a primeira planilha

    times = []  # cria um array vazio
    lin = 2
    for col in xrange(1, 10):
        if ranking.cell(lin, col).value == "Equipe":
            colDeEquipes = col
            break

    lin = 3
    for i in xrange(lin, numeroDeTimesConvidados + lin):
        time = ranking.cell(i, colDeEquipes).value
        time = time.lower().replace(" ", "-").replace(".",
                                                      "-")  # pega o time, passa pra minusculo e tira os caracteres especiais e espacos
        if time[time.__len__() - 1] == '-':  # se o time terminar com "-", remove o "-" da ultima posicao
            time = time[:-1]
        times.append(time)
    return times


# ----------------------------------------------------------------------------------- BUSCA PONTUACOES E SALVA NA PLANILHA
def buscaPontuacoes(workbookPath, dictionary):
    wb = openpyxl.load_workbook(filename=workbookPath)  # abre o arquivo
    ranking = wb.worksheets[0]  # pega a primeira planilha
    pontuacaoSheet = wb.worksheets[2]
    times = []  # cria um array vazio
    lin = 3
    col = 4

    pontuacaoSheet.cell(lin - 1, 1).value = "Time"
    pontuacaoSheet.cell(lin - 1, 2).value = "Pontuacao"
    pontuacaoSheet.cell(lin - 1, 3).value = "Quais ligas passou"

    time = ranking.cell(lin, col).value  # pega o primeiro time do ranking
    while (time != None):  # enquanto o campo nao for vazio
        time = time.lower().replace(" ", "-").replace(".",
                                                      "-")  # pega o time, passa pra minusculo e tira os caracteres especiais e espacos
        if time[time.__len__() - 1] == '-':  # se o time terminar com "-", remove o "-" da ultima posicao
            time = time[:-1]
        times.append(time)
        response = requests.get('https://api.cartolafc.globo.com/time/slug/' + time, verify=False)
        print time + " - " + str(response.json()['pontos']).replace(".", ",")
        # print ".",
        pontuacaoSheet.cell(lin, 1).value = time
        pontuacaoSheet.cell(lin, 2).value = str(response.json()['pontos']).replace(".", ",")
        if time in dictionary:
            pontuacaoSheet.cell(lin, 3).value = dictionary[time]
        else:
            pontuacaoSheet.cell(lin, 3).value = " "
        lin += 1
        time = ranking.cell(lin, col).value  # pega o proximo time do ranking
    # print times
    wb.save(workbookPath)
    print "\nPlanilha salva com sucesso!"
    return times


# ----------------------------------------------------------------------------------- MENU LOGIN
def menuLogin():
    usuario = raw_input(
        "\nBEM VINDO AO MENU DE CONVITES E CONSULTAS DAS LIGAS V.I.C.TROLA\n\nComo quem voce deseja logar?\n[1] - Victor\n[2] - Igor\n[3] - Victor no PC do Igor =)\n")

    if usuario.replace("[", "").replace("]", "") == "1":
        map = ["victorrez85@yahoo.com.br", "",
               "C:\\Users\\victo\\Dropbox\\Outros\\Entretenimento\\Cartola\\Temp2018\\Copa Victrola-2018.xlsx"]
    elif usuario.replace("[", "").replace("]", "") == "2":
        map = ["borges_igor@yahoo.com.br", "",
               'C:\\Users\\igorb\\Documents\\Dropbox\\Temp2017\\Copa Victrola-2018.xlsx']
    elif usuario.replace("[", "").replace("]", "") == "3":
        map = ["victorrez85@yahoo.com.br", "",
               'C:\\Users\\igorb\\Documents\\Dropbox\\Temp2017\\Copa Victrola-2018.xlsx']
    else:
        print "usuario nao encontrado!"
    return map


# ----------------------------------------------------------------------------------- MENU PRINCIPAL
def menuPrincipal():
    menu = raw_input("\n\n------------------------------MENU------------------------------\n"
                     "[1] - Criar Liga\n"
                     "[2] - Convidar pagantes\n"
                     "[3] - Convidar nao pagantes\n"
                     "[4] - Convidar por letra da planilha\n"
                     "[5] - Convidar 5 primeiros do ranking\n"
                     "[6] - Conferir times que ja aceitaram convite\n"
                     "[7] - Conferir times que ainda nao aceitaram o convite\n"
                     "[8] - Busca pontuacao dos times\n"
                     "[9] - Sair\n"
                     "----------------------------------------------------------------\n")
    return menu


# -----------------------------------------------------------------------------------
def main():
    global time
    map = menuLogin()
    token = login(map[0], map[1])

    menu = menuPrincipal()
    ligaCriada = False
    while menu != "9":
        # [1] - Criar Liga
        if menu == "1":
            numeroDaLiga = raw_input("Qual sera a liga (ex: I, II, III ...)?\n")
            liga = numeroDaLiga + " Liga V.I.C.trola"
            quantidadeDeTimes = raw_input("De quantos times sera a liga (4, 8, 16 ou 32)?\n")
            response = criaLiga(token, liga, quantidadeDeTimes)
            if response[0] == 201:
                slug = response[1]
                ligaCriada = True
        # [2] - Convidar pagantes
        elif menu == "2":
            if not ligaCriada:
                numeroDaLiga = raw_input(
                    "Para qual liga deverao ser enviados os convites (ex: I, II, III ...)?\n").lower()
                slug = numeroDaLiga + "-liga-v-i-c-trola"
                ligaCriada = True
            times_para_convidar = buscaTimes("p", map[2], numeroDaLiga)
            print ""
            for time in times_para_convidar:
                print " - " + time
            print ""
            confirmacao = raw_input("Total de " + str(
                times_para_convidar.__len__()) + " times pagantes para convidar.\nConfirma? (s ou n)\n")
            if confirmacao.lower() == "s":
                enviaConvites(token, slug, times_para_convidar)
        # [3] - Convidar nao pagantes
        elif menu == "3":
            if not ligaCriada:
                numeroDaLiga = raw_input(
                    "Para qual liga deverao ser enviados os convites (ex: I, II, III ...)?\n").lower()
                slug = numeroDaLiga + "-liga-v-i-c-trola"
                ligaCriada = True
            times_para_convidar = buscaTimes("f", map[2], numeroDaLiga)
            print "Times nao pagantes encontrados:"
            for time in times_para_convidar:
                print " - " + time
            print ""
            confirmacao = raw_input(
                "Total de " + str(
                    times_para_convidar.__len__()) + " times nao pagantes para convidar.\nConfirma (s ou n)?\n")
            if confirmacao.lower() == "s":
                enviaConvites(token, slug, times_para_convidar)
        # [4] - Convidar por letra da planilha
        elif menu == "4":
            if not ligaCriada:
                numeroDaLiga = raw_input(
                    "Para qual liga deverao ser enviados os convites (ex: I, II, III ...)?\n").lower()
                slug = numeroDaLiga + "-liga-v-i-c-trola"
                ligaCriada = True
            letra = raw_input("Qual a letra da planilha deseja enviar os convites?\n").lower()

            times_para_convidar = buscaTimes(letra, map[2], numeroDaLiga)
            print "Times encontrados:"
            for time in times_para_convidar:
                print " - " + time
            print ""
            confirmacao = raw_input(
                "Total de " + str(
                    times_para_convidar.__len__()) + " times para convidar.\nConfirma (s ou n)?\n")
            if confirmacao.lower() == "s":
                enviaConvites(token, slug, times_para_convidar)
        # [5] - Convidar 5 primeiros do ranking
        elif menu == "5":
            if not ligaCriada:
                numeroDaLiga = raw_input(
                    "Para qual liga deverao ser enviados os convites (ex: I, II, III ...)?\n").lower()
                slug = numeroDaLiga + "-liga-v-i-c-trola"
                ligaCriada = True
            times_para_convidar = buscaTimesPorRanking(5, map[2], numeroDaLiga)
            print "5 primeiros times do ranking"
            for time in times_para_convidar:
                print " - " + time
            print ""
            confirmacao = raw_input(
                "Total de " + str(
                    times_para_convidar.__len__()) + " times (primeiros do ranking).\nConfirma (s ou n)?\n")
            if confirmacao.lower() == "s":
                enviaConvites(token, slug, times_para_convidar)
        # [6] - Conferir times que ja aceitaram convite
        elif menu == "6":
            if not ligaCriada:
                numeroDaLiga = raw_input(
                    "Para qual liga deverao ser enviados os convites (ex: I, II, III ...)?\n").lower()
                slug = numeroDaLiga + "-liga-v-i-c-trola"
                ligaCriada = True
            timesQueJaAceitaram = []
            timesJaConvidados = []
            checaTimesConvidados(token, slug, timesQueJaAceitaram, timesJaConvidados)
            print "Times que ja aceitaram o convite:"
            for time in timesQueJaAceitaram:
                print " - " + time
            print "Total de " + str(timesQueJaAceitaram.__len__()) + " times ja aceitaram o convite"
        # [7] - Conferir times que ainda nao aceitaram o convite
        elif menu == "7":
            if not ligaCriada:
                numeroDaLiga = raw_input(
                    "Para qual liga deverao ser enviados os convites (ex: I, II, III ...)?\n").lower()
                slug = numeroDaLiga + "-liga-v-i-c-trola"
                ligaCriada = True
            timesQueJaAceitaram = []
            timesJaConvidados = []
            checaTimesConvidados(token, slug, timesQueJaAceitaram, timesJaConvidados)
            print "Times que ja foram convidados mas nao aceitaram o convite:"
            for time in timesJaConvidados:
                print " - " + time
            print "Total de " + str(timesJaConvidados.__len__()) + " times nao aceitaram o convite"
        # [8] - Busca pontuacao dos times
        elif menu == "8":
            dictionary = {}
            rodadaAtual = raw_input(
                "Qual o numero da rodada que deseja buscar os confrontos?\n")

            print "Aguarde enquanto buscamos pelas ligas...\n"
            ligas = ['i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x',
                     'xi', 'xii', 'xiii', 'xiv', 'xv', 'xvi', 'xvii', 'xviii', 'xix', 'xx',
                     'xxi', 'xxii', 'xxiii', 'xxiv', 'xxv', 'xxvi', 'xxvii', 'xxviii', 'xxix', 'xxx',
                     'xxxi', 'xxxii', 'xxxiii', 'xxxiv', 'xxxv', 'xxxvi', 'xxxvii', 'xxxviii']
            try:
                for i in ligas:
                    headers = {
                        'x-glb-token': token
                    }
                    ligaMataMata = i + '-liga-v-i-c-trola'
                    response = requests.request("GET", 'https://api.cartolafc.globo.com/auth/liga/' + ligaMataMata,
                                                headers=headers, verify=False)
                    if response.status_code == 404:
                        break
                    for j in response.json()["chaves_mata_mata"][rodadaAtual]:
                        response2 = requests.request("GET",
                                                     'https://api.cartolafc.globo.com/time/id/' + str(j["vencedor_id"]),
                                                     verify=False)
                        # print response2.json()["time"]["slug"]

                        # print response2.json()["time"]["slug"] + "  -  " + str(response2.json()["time"]["slug"] in dictionary)
                        if response2.json()["time"]["slug"] in dictionary:
                            novaLiga = dictionary[response2.json()["time"]["slug"]]
                            dictionary[response2.json()["time"]["slug"]] = (novaLiga + ", " + i)
                            # print dictionary[response2.json()["time"]["slug"]]
                        else:
                            dictionary[response2.json()["time"]["slug"]] = i
                    print ligaMataMata + " - ok"
            except:
                pass
            # for a in dictionary:
            #     print a + " - " + dictionary[a]
            buscaPontuacoes(map[2], dictionary)

        menu = menuPrincipal()


if __name__ == "__main__":
    main()
